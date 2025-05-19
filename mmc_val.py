from openai import OpenAI
from whale import TextGeneration
import time
from datetime import datetime
from pytz import timezone
import pandas as pd
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

apikey = 'XU6BSWEF8I'
TextGeneration.set_api_key(apikey, base_url="https://whale-wave.alibaba-inc.com")


client = OpenAI(
    api_key='sk-144c84438e224beb96b9c10dd9841257',
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)


def generate_image_text_list(url_list):
    # 初始化结果列表
    result = []

    # 遍历 URL 列表，为每个 URL 生成对应的字典
    for url in url_list:
        image_dict = {
            "type": "image_url",
            "image_url": {
                "url": url
            }
        }
        result.append(image_dict)

    return result


def mllm_chat_mmc(url, memo='', cate_name=''):
    content = generate_image_text_list(url)
    # 添加固定的文本部分
    text_dict = {"type": "text", "text":
        f"你是一个经验丰富的商品问题审核专家，负责根据消费者提交的图片、描述（{memo}）、商品品类（{cate_name}），依次完成以下两个任务："
        "任务一：商品品类一致性校验"
        f"1. 分析图片内容：仔细查看消费者上传的图片，结合消费者描述（{memo}），确定图片中“实际品类”。"
        "注意当图片中有明显的买家和卖家/平台协商记录信息时，则图片中“实际品类”校正为“无效url”，商品品类一致性校验不通过返回false"
        f"2. 一致性校验：判断图片、描述（{memo}）中的“实际品类”与商品品类（{cate_name}）是否一致。"
        f"注意，“实际品类”不需要过度细化，只要能和商品品类（{cate_name}）相对应即为一致，例如识别出香梨和商品品类（梨）是一致的"

        "任务二：商品问题识别"
        f"1. 分析消费者上传的图片推断商品存在的实际问题："
        "仔细查看消费者上传的图片，从图片中已发生的事实来判断商品的实际问题"
        "必须要针对图片中已经发生且真实存在的问题进行判断，不要进行任何推测（例如可能导致发霉腐烂、不确定内部商品安全这种推测结果）"
        "只针对图片中存在的实际问题进行独立的问题识别，不要将问题识别和商品品类一致性校验耦合在一起，禁止将“非商品品类”作为问题结果。"
        
        "2. 从下面提供的这几种问题结果中选择和商品实际问题相近的那一种作为最终的商品问题。"
        "   - 包装/商品破损/污渍:物流或外力造成的快递包装物理性损坏例如快递包装的破损、脏污、开裂、变形（如快递纸箱撕裂、快递箱存在明显水渍污渍等）、商品包装或容器（如盛装蓝莓的塑料盒等）破损变形、商品自身或商品内容物（例如内含的调料包）破损挤碎变形（如面饼碎裂、瓜果被挤压导致的开裂掉块）、漏液洒出（酱油瓶破损导致酱油洒出等）"
        "   - 商品变质:指商品自身出现的临期过期、冷冻商品化冻（本应在低温环境存储运输的商品，发现冷冻状态失效出现化冻、融化等问题）、新鲜度不足、明显异物、霉变或虫害引起的黑斑、发霉、变质腐烂。"
        "注意：“包装/商品破损/污渍”包含了全部的破损问题，例如快递包装破损、商品外包装破损、商品自身破损等，只要是破损问题都归到“包装/商品破损/污渍”。"
        
        "### 注意事项"
        "每个任务都请务必逐步分析、仔细查看图片、描述等内容，确保判断准确。"
        "你的所有判断都必须基于已发生的事实给出。"
        
        "必须严格按照以下格式以JSON返回结果，不需要返回其他内容，用中文回答"
        "cateCheckReason: 商品品类一致性校验的依据。"
        "qualityCheckReason: 判商品是否存在问题的原因描述，简单原因描述限制在20字以内。"
        "cateCheckResult: true或false，表示消费者上传图片对应的实际品类和商品品类是否一致。"
        "qualityCheckResult: true或false，表示消费者上传图文是否存在实际的商品问题，不存在任何已发生的实际问题则为false。"
        "qualityCheckScene: 最终商品问题，如果`qualityCheckResult`是false则`qualityCheckScene`返回“无明显问题”。"
        "示例："
        "{"
        "“cateCheckReason”: “xxxxx”,"
        "“qualityCheckReason”: “xxxxx”,"
        "“cateCheckResult”: “xxx”,"
        "“qualityCheckResult”: “xxx”,"
        "“qualityCheckScene”: “xxx”"
        "}"


    }
    content.append(text_dict)
    try:
        completion = client.chat.completions.create(
            temperature=0.0,
            model="qwen-vl-max",
            messages=[
                {
                    "role": "system",
                    "content": [{"type": "text",
                                 "text": "You are a good assistant"}],
                },
                {
                    "role": "user",
                    "content": content,
                },
            ],
        )
        return completion.choices[0].message.content
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return 'Error'


def request_whale(url):
    content = generate_image_text_list(url)
    # 添加固定的文本部分
    text_dict = {"type": "text", "text":"你是一个商品问题审核员，请你结合所给图片判断出现的商品问题、影响程度以及问题商品个数"}
    content.append(text_dict)
    try:
        msgs = [
            {
              "role": "system",
              "content": "You are a helpful assistant."
            },
            {
              "role": "user",
              "content": content
            }
          ]

        extend_fields= {"top_k": 1}

        # 请求模型
        response = TextGeneration.chat(
            model="proof_maochao_xiaomi_test",
            messages=msgs,
            stream=False,
            temperature=0.1,
            timeout=60,
            max_tokens=4096,
            top_p=0.8,
            extend_fields=extend_fields)

        # 处理结果
        # print(json.dumps(response.extend_fields, ensure_ascii=False))
        # print(json.dumps(response.choices[0].message.content, ensure_ascii=False))
        return response.choices[0].message.content
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return 'Error'


def llm_chat_mmc(cate_name,sku,itemBuyAmount=1,itemTitle=''):
    try:
        completion = client.chat.completions.create(
            model="qwen2.5-72b-instruct",
            messages=[
                {'role': 'system', 'content': 'You are a helpful assistant.'},
                {'role': 'user', 'content':
                    '###角色'
                    '你是一个商品问题审核员，擅长根据用户购买的商品类目、规格信息和购买件数计算总件数。'
                    '###任务'
                    '根据商品品类、规格信息和购买件数计算总件数。你需要逐步分析并计算，确保每一步逻辑清晰、准确。'
                    '###计算规则'
                    '1. 包装规格明确的情况（如：瓶、盒等）'
                    '如果规格信息中明确说明了单件内的数量（如“2瓶”），则直接用该数量乘以购买件数得出总件数。'
                    '如果规格信息中说明给定个数范围且极差不超过1则取最大值乘以购买件数得出总件数，否则直接先取均值再乘以购买件数得出总件数。'
                    # '示例：'
                    # '商品品类：“白酒/调香白酒”，规格信息：“规格描述:1瓶”，购买件数：“3”，则总件数为1瓶*3=3。'
                    # '商品品类：“百香果”，规格信息：“水果总重:5斤【6-10个】;单果规格:中果”，购买件数：“1”，则总件数为平均值8个*1=8。'
                    
                    '2. 总斤数和果径的情况'
                    '如果规格信息中提供了果径范围和总重量，则按照以下步骤推测单件内的个数：'
                    'Step 1 : 确认商品的品类。'
                    'Step 2 : 根据商品品类和果径范围推测单果重量（单位：克）。例如，直径80mm-85mm的苹果单果重量约为200g-300g，取平均值250g。'
                    'Step 3 : 将总重量转换为克（1斤=500g），然后除以单果重量得到单件内的个数。'
                    'Step 4 : 将单件内的个数乘以购买件数得出总件数。'
                    # '示例：'
                    # '商品品类：“苹果”，规格信息：“苹果果径:80mm-85mm;重量:5斤”，购买件数：“2”'
                    # 'Step 1: 确认商品的品类为苹果。'
                    # 'Step 2: 苹果在果径:80mm-85mm的情况下，单果重量约为200g-300g，取平均值250g。'
                    # 'Step 3: 5斤=2500g，2500 / 250 = 10个。'
                    # 'Step 4: 总件数=10 * 2 = 20个。'

                    '3. 总斤数和单果重的情况'
                    '如果规格信息中提供了总重量和单果重量范围，则按照以下步骤推测单件内的个数：'
                    'Step 1 : 确认商品的品类。'
                    'Step 2 : 明确单果重量，如果直接给出单果200g这种准确重量则直接使用，如果给出60-90g这种重量范围则取均值作为单果重量'
                    'Step 3 : 将总重量转换为克（1斤=500g），然后除以单果重量得到单件内的个数。'
                    'Step 4 : 将单件内的个数乘以购买件数得出总件数。'
                    # '示例：'
                    # '商品品类：“柠檬”，规格信息：“单果重量:60-90g;重量:5斤”，购买件数：“1”'
                    # 'Step 1: 确认商品为柠檬。'
                    # 'Step 2: 单果重量:60-90g，重量是个范围因此单果重量取均值约75g'
                    # 'Step 3: 5斤=2500g，2500 / 75 = 33个'
                    # 'Step 4: 总件数=33 * 1 = 33个'
                    
                    '4. 总斤数和大中小果的情况'
                    '如果规格信息中仅说明总重量和大中小果规格，则按照以下步骤推测单件内的个数：'
                    'Step 1 : 确认商品的品类。'
                    'Step 2 : 根据对商品品类的经验预估单果重量。例如，红薯的大果约为300g，中果约为200g，小果约为100g。'
                    'Step 3 : 将总重量转换为克，然后除以单果重量得到单件内的个数。'
                    'Step 4 : 将单件内的个数乘以购买件数得出总件数。'
                    # '示例：'
                    # '商品品类：“番薯”，规格信息：“重量:9斤;单果规格:精选大果”，购买件数：“1”'
                    # 'Step 1: 确认商品品类为红薯。'
                    # 'Step 2: 根据对红薯的经验，精选大果红薯单果重量约为300g。'
                    # 'Step 3: 9斤=4500g，4500 / 300 = 15个。'
                    # 'Step 4: 总件数=15 * 1 = 15个。'
                    
                    '5. 其他情况'
                    '如果规格信息中只有总重且没有任何单果信息，则认为只含一件，1*购买件数=总件数'
                    # '示例：'
                    # '商品品类：“甜瓜”，规格信息：“重量:4.5-5斤”,购买件数：“1”'
                    # '只有总重量且没有任何单果信息则认为只含一件，1*购买件数=购买件数'
                    
                    
                    
                    # f'如果规格信息中只有总重且没有任何单果信息，则需根据商品标题 {itemTitle} 分析实际品类，并按以下规则处理：'
                    # '第一步 : 确认商品的实际品类及其具体品种。'
                    # '如果商品标题中包含具体的品种名称（如“麒麟西瓜”、“冰糖子西瓜”），则需根据品种进一步判断其果径大小。'
                    # '如果商品标题中未明确品种，则默认按照品类的一般特性处理。'
                    # '第二步 : 判断该具体商品品种是否属于超大果径品类。'
                    # '1. 超大果径品类 包括但不限于：西瓜、榴莲、菠萝蜜、南瓜等。'
                    # '对于超大果径品类（如麒麟西瓜、榴莲、菠萝蜜等），无论总重量多少，均认为单件内只含一件，总件数 = 1 * 购买件数。'
                    # '2. 中小果径品类 包括但不限于：苹果、土豆、甜瓜、冰糖子西瓜等。'
                    # '对于中小果径品类（如苹果、土豆、甜瓜等），统一认为是中果，并按照“四. 总斤数和大中小果的情况”计算总件数。'
                    # 
                    # '示例：'
                    # "商品品类：“香蕉”，规格信息：“水果总重:整箱5-6斤”，购买件数：“2件”"
                    # 'Step 1: 规格信息中没有任何单果信息，确认属于“五. 其他情况”。'
                    # 'Step 2: 因为命中了“五. 其他情况”，因此参考商品标题：“云南香蕉5斤精选装当季水果青皮蕉自行催熟鲜嫩越甜整箱包邮发货”属于小果径商品品类。'
                    # 'Step 3: 确认是小果径商品，因此按照“四. 总斤数和大中小果的情况”，以中果规格来计算总件数。'
                    # 
                    # '示例：'
                    # '商品品类：“菠萝蜜”，规格信息：“重量:0斤;单果重量:25-30斤【严选山地老树果爆甜】”，购买件数：“1”'
                    # 'Step 1: 规格信息中没有任何单果信息，确认属于“五. 其他情况”。'
                    # 'Step 2:  因为命中了“五. 其他情况”，因此参考商品标题：“正宗海南菠萝蜜一整个新鲜水果脆甜当季现摘干苞黄肉整箱包邮”属于大果径商品品类。'
                    # 'Step 2: 根据商品标题：“正宗海南菠萝蜜一整个新鲜水果脆甜当季现摘干苞黄肉整箱包邮”判断出是超大果径的商品，因此总件数 = 1 * 1 = 1个。'
                    
                    f'商品品类：“{cate_name}”，规格信息：“{sku}”，购买件数：“{itemBuyAmount}”'
                    '###限制'
                    # '只有在规格信息中没有任何单果信息命中“五. 其他情况”时才可以使用商品标题信息，命中“包装规格明确”、“总斤数和果径”、“总斤数和单果重”、“总斤数和大中小果”时都禁止使用标题信息。'
                    '仅以json格式返回最终结果，用中文回答'
                    'reason:返回你的分析过程'
                    'branch:规格信息能提取到的情况（包装规格明确、总斤数和果径、总斤数和单果重、总斤数和大中小果、其他情况）'
                    'totalBuy:总个数'
                    # '示例：'
                    # '{'
                    #   '“reason”: “商品品类为百香果，规格信息显示总重量为3斤，单果规格为大果60-70g。首先确认商品品类为百香果，单果重量范围为60-70g，取平均值65g。将总重量3斤转换为克，即1500g。然后用总重量1500g除以单果重量65g，得到单件内的个数约为23个（四舍五入）。因为购买件数为1，所以总件数为23个。”,'
                    #   '“branch”: “总斤数和单果重”'
                    #   '“totalBuy”: “23”'
                    # '}'
                }],
        )
        return completion.choices[0].message.content
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return 'Error'

"""
if __name__ == "__main__":
    # url = ["https://img.alicdn.com/imgextra/i2/2218399989887/O1CN01dno3ki2MuIwtTbCcn_!!2218399989887-0-refund_platform.jpg", "https://img.alicdn.com/imgextra/i3/2218399989887/O1CN019Q2hCU2MuIwsy9TtP_!!2218399989887-0-refund_platform.jpg"]
    # cate_name = "生猪肉"
    # memo = ""
    # titlesku = "蚕蛹鲜活新鲜金丝蛹 山东特产桑蚕蛹子生茧蝉蛹冷冻 批发价年货,口味:2斤"

    # url = ["https://img.alicdn.com/imgextra/i3/47172280/O1CN011SK3z11SiI5Eq0pJ4_!!47172280-0-xiaomi.jpg","https://img.alicdn.com/imgextra/i1/47172280/O1CN01G18jz61SiI5GZUDNF_!!47172280-0-xiaomi.jpg"]
    #
    # memo = "凹进去了"
    # cate_name = "厨房/烹饪用具>>烧烤/烘焙用具>>烧烤架炉/烧烤DIY用具>>锡纸/油纸"
    # title = '空气炸锅专用纸锡纸盘盒烧烤箱烘焙锡箔碗家用硅吸油纸食品级食物'
    # sku = ''

    url = [
        "https://picasso-work.alibaba-inc.com/i2/O1CN01jTms7O1lSEtnfQVDq_!!6000000004817-0-tps_intranet-800-800.jpg"
    ]
    cate_name = "饮料"
    itemTitle = '测试品'
    sku = "测试品sku 10L/罐"
    itemBuyAmount = 1
    memo = ""

    # res1 = mllm_chat_mmc(url, memo, cate_name)
    # res2 = request_whale(url)
    res3 = llm_chat_mmc(cate_name, sku, itemBuyAmount, itemTitle)
    # print(res1)
    # print(res2)
    print(res3)

#"""


if __name__ == "__main__":
    input_excel = "val.xlsx"
    output_file = "val_res.xlsx"
    separator = ';'
    df = pd.read_excel(input_excel)
    # val_df = df[df['url_cnt'] == 1].drop_duplicates()
    val_df = df[df['lst_reason_name'].isin(["包装/商品破损/污渍", "商品变质"])].drop_duplicates()
    # val_df = val_df1[val_df1['fst_refund_desc'] != '\\N']
    print(val_df.shape)
    for index, row in val_df.iterrows():
        # 获取当前时间戳
        timestamp = time.time()

        # 将时间戳转换为 UTC 时间
        utc_time = datetime.utcfromtimestamp(timestamp)

        # 设置北京时区
        beijing_tz = timezone('Asia/Shanghai')

        # 将 UTC 时间转换为北京时间
        beijing_time = utc_time.replace(tzinfo=timezone('UTC')).astimezone(beijing_tz)

        # 格式化输出北京时间
        formatted_beijing_time = beijing_time.strftime('%Y-%m-%d %H:%M:%S')

        print(f'第{index}条数据开始处理时间：{formatted_beijing_time}')


        url = row['actionkeypics'].split(',')
        cate_name = row['cate_name']
        itemTitle = row['item_title']
        sku = row['extracted_sku_info'].split('|')[-1].replace('#3B', ':').replace('#3A', ';')
        memo = row['fst_refund_desc']
        itemBuyAmount = row['itemBuyAmount']

        start_time = time.time()
        res1 = mllm_chat_mmc(url, memo, cate_name)
        end_time = time.time()
        # 计算运行时间
        print(f"Qwen-vl-max代码运行时间: {end_time - start_time:.6f} 秒")
        cost_time1 = end_time - start_time
        if res1 == 'Error':
            val_df.loc[index, 'res1_cateCheckResult'] = 'Error'
            val_df.loc[index, 'res1_qualityCheckResult'] = 'Error'
            val_df.loc[index, 'res1_qualityCheckScene'] = 'Error'
        else:
            res1_str = res1.split('{')[-1].split('}')[0].split(',')
            val_df.loc[index, 'res1_cateCheckResult'] = res1_str[0].split(':')[-1]
            val_df.loc[index, 'res1_qualityCheckResult'] = res1_str[2].split(':')[-1]
            val_df.loc[index, 'res1_qualityCheckScene'] = res1_str[3].split(':')[-1]

        start_time = time.time()
        res2 = request_whale(url)
        end_time = time.time()
        # 计算运行时间
        print(f"whale代码运行时间: {end_time - start_time:.6f} 秒")
        cost_time2 = end_time - start_time
        if '无法判断' in res2:
            val_df.loc[index, 'res2_damageCnt'] = '图像无法判断存在质量问题'
        elif res2 == 'Error':
            val_df.loc[index, 'res2_damageCnt'] = 'Error'
        else:
            res2_str = res2.split('{')[-1].split('}')[0].split(',')
            val_df.loc[index, 'res2_damageCnt'] = res2_str[-1].split(':')[-1]

        start_time = time.time()
        res3 = llm_chat_mmc(cate_name, sku, itemBuyAmount, itemTitle)
        end_time = time.time()
        # 计算运行时间
        print(f"Qwen2.5-72b代码运行时间: {end_time - start_time:.6f} 秒")
        cost_time3 = end_time - start_time

        if res3 == 'Error':
            val_df.loc[index, 'res3_totalBuy'] = 'Error'
        else:
            res3_str = res3.split('{')[-1].split('}')[0].split(',')
            val_df.loc[index, 'res3_totalBuy'] = res3_str[0].split(':')[-1]

        val_df.loc[index, 'cost_time1'] = round(cost_time1, 3)
        val_df.loc[index, 'cost_time2'] = round(cost_time2, 3)
        val_df.loc[index, 'cost_time3'] = round(cost_time3, 3)

        try:
            # 如果文件已存在，则加载现有工作簿
            book = load_workbook(output_file)

            # 获取目标工作表的最大行号
            sheet = book['Sheet1']
            start_row = sheet.max_row

            # 使用 openpyxl 直接写入数据
            for col_num, value in enumerate(val_df.loc[index].values, start=1):
                sheet.cell(row=start_row + 1, column=col_num, value=value)

            # 保存工作簿
            book.save(output_file)

        except FileNotFoundError:
            # 如果文件不存在，则创建新文件并写入数据
            temp_df = val_df.loc[[index]]  # 取出当前行
            temp_df.to_excel(output_file, index=False, sheet_name='Sheet1')
